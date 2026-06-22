"""
Sensor Tower 数据 - 上线首年模式.

对一份 ST 导出做以下流水线:
  1. 按 launches.json 里每款游戏的「上线月」截取 (launch_month + 11 个月 = 12 月) 数据
  2. 应用 sensor-tower-data-adjust 口径 (CN ×4/×2.2/×3, 其他 ÷0.7, 详见 build.py)
  3. 输出 Excel: 明细 / 按App汇总 / 按地区汇总. 「按App汇总」附加 Months_Covered + Incomplete 列;
     Incomplete=Y 的行整行染红.

「联网核验上线日期」这一步由 Claude 在调用前完成 (用 WebSearch + WebFetch 查权威源,
呈现 Markdown 表格给用户二次确认), 不在脚本范围内. 用户确认的结果以 JSON 文件传入.

用法:
    python3 first_year_workflow.py <input_csv> <launches_json> [<output_xlsx>]

launches_json 示例 (UTF-8):
    {
      "window_months": 12,
      "launches": {
        "Genshin Impact": "2020-09-28",
        "Honkai: Star Rail": "2023-04-26"
      }
    }

- key 必须是 CSV 里 Unified Name 的精确字符串.
- value 是「全球上线日期」字符串, 可带日, 脚本自动 floor 到月初.
- window_months 可选, 默认 12.
- launches 里没列的游戏会被丢弃 (打 warn).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Dict, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 复用 build.py 里的逻辑
sys.path.insert(0, str(Path(__file__).parent))
from build import adjust, find_col, read_sensor_tower, summarize_by, write_excel  # noqa: E402


def load_config(path: Path) -> Tuple[Dict[str, str], int]:
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    launches = cfg.get("launches") or cfg
    if not isinstance(launches, dict) or not launches:
        raise ValueError(f"{path} 必须含 launches dict (game name -> launch date)")
    window = int(cfg.get("window_months", 12))
    return launches, window


def filter_first_year(
    df: pd.DataFrame, name_col: str, date_col: str, launches: Dict[str, str], window: int
) -> Tuple[pd.DataFrame, Dict[str, dict]]:
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col])

    pieces = []
    meta: Dict[str, dict] = {}

    for name, raw_date in launches.items():
        start = pd.to_datetime(raw_date).to_period("M").to_timestamp()
        end = start + pd.DateOffset(months=window - 1)

        game_df = df[df[name_col] == name]
        if len(game_df) == 0:
            print(f"[warn] '{name}': 0 rows in source — skip")
            meta[name] = {"months_covered": 0, "incomplete": "Y", "last_month": "-"}
            continue

        win = game_df[(game_df[date_col] >= start) & (game_df[date_col] <= end)]
        months_covered = win[date_col].nunique()
        incomplete = "Y" if months_covered < window else "N"
        last_month = win[date_col].max().strftime("%Y-%m") if len(win) else "-"

        meta[name] = {
            "launch_month": start.strftime("%Y-%m"),
            "months_covered": months_covered,
            "incomplete": incomplete,
            "last_month": last_month,
            "rows": len(win),
        }

        pieces.append(win)
        flag = "  [INCOMPLETE]" if incomplete == "Y" else ""
        print(
            f"[filter] {name}: months={months_covered}/{window}  rows={len(win):,}  "
            f"({start.strftime('%Y-%m')} ~ {last_month}){flag}"
        )

    # 未在 launches 里出现的游戏
    listed = set(launches.keys())
    leftover = sorted(set(df[name_col].dropna().unique()) - listed)
    for n in leftover:
        print(f"[skip] '{n}' not in launches.json — dropped")

    if not pieces:
        raise ValueError("过滤后无数据。检查 launches.json 里的 Unified Name 是否与 CSV 一致")

    filtered = pd.concat(pieces, ignore_index=True)
    filtered[date_col] = filtered[date_col].dt.strftime("%Y-%m-%d")
    return filtered, meta


def annotate_excel(dst: Path, meta: Dict[str, dict], name_col: str) -> None:
    """给「按App汇总」sheet 追加 Months_Covered / Incomplete 列, Incomplete=Y 整行染红。"""
    wb = load_workbook(dst)
    ws = wb["按App汇总"]
    header = [c.value for c in ws[1]]
    name_idx = header.index(name_col) + 1
    next_col = len(header) + 1

    ws.cell(row=1, column=next_col, value="Months_Covered").font = Font(bold=True)
    ws.cell(row=1, column=next_col + 1, value="Incomplete").font = Font(bold=True)

    red = Font(color="C00000", bold=True)
    incomp_letter = get_column_letter(next_col + 1)

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_idx).value
        info = meta.get(name)
        if not info:
            ws.cell(row=row, column=next_col, value="?")
            ws.cell(row=row, column=next_col + 1, value="?")
            continue
        ws.cell(row=row, column=next_col, value=info["months_covered"]).number_format = "0"
        ws.cell(row=row, column=next_col + 1, value=info["incomplete"])
        if info["incomplete"] == "Y":
            for col in range(1, next_col + 2):
                ws.cell(row=row, column=col).font = red
            for cell in ws[incomp_letter][1:]:
                if cell.value == "Y":
                    cell.font = red

    wb.save(dst)


def main(argv: list) -> None:
    if len(argv) < 3:
        print(__doc__)
        sys.exit(1)

    src = Path(argv[1]).expanduser().resolve()
    cfg_path = Path(argv[2]).expanduser().resolve()
    if not src.exists():
        raise FileNotFoundError(src)
    if not cfg_path.exists():
        raise FileNotFoundError(cfg_path)
    dst = Path(argv[3]).expanduser().resolve() if len(argv) >= 4 else src.with_name(src.stem + "_首年_adjusted.xlsx")

    launches, window = load_config(cfg_path)
    print(f"[config] {len(launches)} games, window={window} months  ←  {cfg_path.name}")

    df = read_sensor_tower(src)
    print(f"[load] {len(df):,} rows  ←  {src.name}")

    name_col = find_col(df, "Unified Name", "App Name", "Name") or df.columns[0]
    date_col = find_col(df, "Date") or "Date"

    filtered, meta = filter_first_year(df, name_col, date_col, launches, window)
    print(f"[filter] kept {len(filtered):,} rows ({len([m for m in meta.values() if m.get('rows', 0) > 0])} games)")

    filtered, adj_meta = adjust(filtered)

    by_app = summarize_by(filtered, name_col)
    by_region = summarize_by(filtered, adj_meta["country_col"])

    write_excel(filtered, by_app, by_region, dst)
    annotate_excel(dst, meta, name_col)

    print()
    print("=" * 60)
    print(f"[done] 写入: {dst}")
    print()
    print("[summary]")
    summary = pd.DataFrame(
        [
            {
                "Unified Name": name,
                "Launch Month": info.get("launch_month", "?"),
                "Months_Covered": info["months_covered"],
                "Last Month": info.get("last_month", "?"),
                "Rows": info.get("rows", 0),
                "Incomplete": info["incomplete"],
            }
            for name, info in meta.items()
        ]
    )
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main(sys.argv)
