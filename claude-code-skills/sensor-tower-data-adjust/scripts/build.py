"""
Sensor Tower 源数据投资口径调整。

调整规则:
- Country/Region == 'CN':
    Downloads_adj   = Downloads * 4
    Revenue_adj     = Revenue / 0.7 * 2.2
    DAU_adj         = DAU * 3                  (DAU 列存在时)
- 其他地区:
    Downloads_adj   = Downloads
    Revenue_adj     = Revenue / 0.7
    DAU_adj         = DAU                       (DAU 列存在时)
- RPD_adj           = Revenue_adj / Downloads_adj
- ARPDAU_adj        = Revenue_adj / DAU_adj    (DAU 列存在时)

用法:
    python3 build.py <input_csv> [<output_xlsx>]

输入自动识别 UTF-16 / UTF-8 编码和 tab / comma 分隔符。
输出 Excel 3 张 sheet: 明细 / 按App汇总 / 按地区汇总。
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

# ----- 口径常量（如需调整改这里）-----
CN_CODES = {"CN"}            # 视为"中国"的 Country/Region 代码
PLATFORM_SHARE = 0.7         # Apple/Google 平台分成 → 还原乘数 = 1/0.7
CN_REV_MULT = 2.2            # CN 区收入额外上修
CN_DL_MULT = 4               # CN 区下载量上修
CN_DAU_MULT = 3              # CN 区 DAU 上修
# ------------------------------------


def read_sensor_tower(path: Path) -> pd.DataFrame:
    """自动识别编码和分隔符读 ST 导出文件。"""
    last_err: Optional[Exception] = None
    for encoding in ("utf-16", "utf-8-sig", "utf-8"):
        for sep in ("\t", ","):
            try:
                df = pd.read_csv(path, encoding=encoding, sep=sep)
                if df.shape[1] >= 5:
                    sep_label = "TAB" if sep == "\t" else "COMMA"
                    print(f"[load] encoding={encoding} sep={sep_label}")
                    return df
            except Exception as e:
                last_err = e
                continue
    raise RuntimeError(f"无法解析文件 {path}: {last_err}")


def find_col(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    """找列名，容忍空格 / 大小写 / 下划线差异。"""
    norm = {c.strip().lower().replace(" ", "").replace("_", ""): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower().replace(" ", "").replace("_", "")
        if key in norm:
            return norm[key]
    return None


def resolve_paths(argv: list) -> Tuple[Path, Path]:
    if len(argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = Path(argv[1]).expanduser().resolve()
    if not src.exists():
        raise FileNotFoundError(src)
    if len(argv) >= 3:
        dst = Path(argv[2]).expanduser().resolve()
    else:
        dst = src.with_name(src.stem + "_adjusted.xlsx")
    return src, dst


def adjust(df: pd.DataFrame) -> Tuple[pd.DataFrame, dict]:
    country_col = find_col(df, "Country / Region", "Country/Region", "Country", "Country Code")
    downloads_col = find_col(df, "Downloads", "Download")
    revenue_col = find_col(df, "Revenue ($)", "Revenue($)", "Revenue", "Revenue (USD)")
    dau_col = find_col(df, "DAU", "Daily Active Users", "Active Users")
    name_col = find_col(df, "Unified Name", "App Name", "Name") or df.columns[0]

    missing = [n for n, c in [("country", country_col), ("downloads", downloads_col), ("revenue", revenue_col)] if c is None]
    if missing:
        raise ValueError(f"必需列缺失: {missing}\n实际列: {list(df.columns)}")

    print(f"[map] country='{country_col}' downloads='{downloads_col}' revenue='{revenue_col}'"
          f" dau={'<%s>' % dau_col if dau_col else 'none'} name='{name_col}'")

    is_cn = df[country_col].isin(CN_CODES)
    df["Is_CN"] = np.where(is_cn, "Y", "N")
    df["Downloads_adj"] = np.where(is_cn, df[downloads_col] * CN_DL_MULT, df[downloads_col])
    df["Revenue_adj ($)"] = np.where(
        is_cn,
        df[revenue_col] / PLATFORM_SHARE * CN_REV_MULT,
        df[revenue_col] / PLATFORM_SHARE,
    )
    with np.errstate(divide="ignore", invalid="ignore"):
        df["RPD_adj ($)"] = np.where(
            df["Downloads_adj"] > 0, df["Revenue_adj ($)"] / df["Downloads_adj"], np.nan
        )

    has_dau = dau_col is not None
    if has_dau:
        df["DAU_adj"] = np.where(is_cn, df[dau_col] * CN_DAU_MULT, df[dau_col])
        with np.errstate(divide="ignore", invalid="ignore"):
            df["ARPDAU_adj ($)"] = np.where(
                df["DAU_adj"] > 0, df["Revenue_adj ($)"] / df["DAU_adj"], np.nan
            )

    meta = {
        "country_col": country_col,
        "downloads_col": downloads_col,
        "revenue_col": revenue_col,
        "dau_col": dau_col,
        "name_col": name_col,
        "has_dau": has_dau,
        "is_cn_mask": is_cn,
    }
    return df, meta


def summarize_by(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    agg = (
        df.groupby(group_col, dropna=False)
        .agg(
            Downloads_adj=("Downloads_adj", "sum"),
            **{"Revenue_adj ($)": ("Revenue_adj ($)", "sum")},
            Rows=("Downloads_adj", "size"),
        )
        .reset_index()
    )
    agg["RPD_blended ($)"] = np.where(
        agg["Downloads_adj"] > 0, agg["Revenue_adj ($)"] / agg["Downloads_adj"], np.nan
    )
    return agg.sort_values("Revenue_adj ($)", ascending=False).reset_index(drop=True)


def write_excel(detail: pd.DataFrame, by_app: pd.DataFrame, by_region: pd.DataFrame, dst: Path) -> None:
    int_fmt = "#,##0"
    money_fmt = "#,##0.00"
    rate_fmt = "#,##0.0000"
    col_fmts_by_sheet = {
        "明细": {
            "Downloads": int_fmt,
            "Downloads_adj": int_fmt,
            "Revenue ($)": money_fmt,
            "Revenue_adj ($)": money_fmt,
            "RPD ($)": rate_fmt,
            "RPD_adj ($)": rate_fmt,
            "DAU": int_fmt,
            "DAU_adj": int_fmt,
            "ARPDAU_adj ($)": rate_fmt,
        },
        "按App汇总": {
            "Downloads_adj": int_fmt,
            "Revenue_adj ($)": money_fmt,
            "RPD_blended ($)": rate_fmt,
            "Rows": int_fmt,
        },
        "按地区汇总": {
            "Downloads_adj": int_fmt,
            "Revenue_adj ($)": money_fmt,
            "RPD_blended ($)": rate_fmt,
            "Rows": int_fmt,
        },
    }

    with pd.ExcelWriter(dst, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="明细", index=False)
        by_app.to_excel(writer, sheet_name="按App汇总", index=False)
        by_region.to_excel(writer, sheet_name="按地区汇总", index=False)
        wb = writer.book
        for sheet_name, col_fmts in col_fmts_by_sheet.items():
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            for col_name, fmt in col_fmts.items():
                if col_name in header:
                    letter = get_column_letter(header.index(col_name) + 1)
                    for cell in ws[letter][1:]:
                        cell.number_format = fmt
            ws.freeze_panes = "A2"


def main(argv: list) -> None:
    src, dst = resolve_paths(argv)
    df = read_sensor_tower(src)
    print(f"[load] {len(df):,} rows, {len(df.columns)} cols  ←  {src}")

    df, meta = adjust(df)

    by_app = summarize_by(df, meta["name_col"])
    by_region = summarize_by(df, meta["country_col"])

    write_excel(df, by_app, by_region, dst)

    print()
    print("=" * 60)
    print(f"[done] 写入: {dst}")
    print(f"[qa] 总行数: {len(df):,}")
    print(f"[qa] DAU/ARPDAU: {'已生成' if meta['has_dau'] else '未生成（无 DAU 列）'}")
    cn_rows = int(meta["is_cn_mask"].sum())
    print(f"[qa] CN 行数: {cn_rows}")
    print()
    print("[qa] 按地区:")
    print(by_region.to_string(index=False))
    print()
    print("[qa] Top 5 App (按 Revenue_adj):")
    print(by_app.head(5).to_string(index=False))
    print()
    rev_orig = df[meta["revenue_col"]].sum()
    rev_adj = df["Revenue_adj ($)"].sum()
    dl_orig = df[meta["downloads_col"]].sum()
    dl_adj = df["Downloads_adj"].sum()
    print(f"[qa] Revenue   原始={rev_orig:,.2f}  调整后={rev_adj:,.2f}  比率={rev_adj/rev_orig:.4f}")
    print(f"[qa] Downloads 原始={dl_orig:,.0f}  调整后={dl_adj:,.0f}  比率={dl_adj/dl_orig:.4f}")
    print()
    print("[hint] 无 CN 时 Revenue 比率应 ≈ 1.4286 (=1/0.7), Downloads 比率应 = 1.0")


if __name__ == "__main__":
    main(sys.argv)
